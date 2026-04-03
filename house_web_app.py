from __future__ import annotations
import streamlit as st
import pandas as pd
import numpy as np
import os
import re
import html
import warnings
import joblib
from io import BytesIO
from datetime import datetime
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

warnings.filterwarnings("ignore")

# Đường dẫn file mẫu và dữ liệu
TEMPLATE_PATH = "data/processed/house_template.csv"
DATA_PATH = "data/processed/non_outlier_samples_6cols.csv"
UPLOAD_DIR = "data/processed/"
LIST_PAGE_SIZE = 20
ADMIN_PATH = "data/processed/admin_new_houses.csv"
SEGMENT_MODEL_DIR = "models/kmeans_by_district"
SEGMENT_FEATURE_COLS = ["so_phong_ngu", "dien_tich_num", "chieu_ngang_num", "log_gia_ban"]
DEFAULT_CLUSTER_LABEL_MAP = {
    0: "nhà cao cấp",
    1: "nhà cận cao cấp",
    2: "nhà bình dân",
}
EXCEL_REQUIRED_EXTRA_COLS = [
    "chieu_ngang_num",
    "chieu_dai_num",
    "log_gia_ban",
]


# =============================
# Global UI styling
# =============================

def inject_global_css() -> None:
    """Tiêm CSS để giao diện hiện đại và đồng bộ hơn."""
    st.markdown(
        """
        <style>
        .main .block-container {
            padding-top: 2rem;
            padding-bottom: 2rem;
            max-width: 1100px;
        }

        .section-title {
            font-size: 1.4rem;
            font-weight: 700;
            margin-bottom: 0.75rem;
        }

        .house-card {
            border-radius: 12px;
            padding: 1rem 1.4rem;
            margin-bottom: 0.9rem;
            background: linear-gradient(135deg, #ffffff, #f5f7ff);
            border: 1px solid #e5e7eb;
            box-shadow: 0 4px 14px rgba(15, 23, 42, 0.06);
        }

        .house-card-title {
            font-weight: 700;
            font-size: 1.0rem;
            margin-bottom: 0.25rem;
        }

        .house-card-meta {
            color: #6b7280;
            font-size: 0.9rem;
            margin-bottom: 0.35rem;
        }

        .house-price {
            color: #16a34a;
            font-weight: 700;
            font-size: 0.95rem;
        }

        .seller-hero {
            border-radius: 14px;
            padding: 1rem 1.2rem;
            margin: 0.3rem 0 1rem 0;
            background: linear-gradient(135deg, #eff6ff, #ffffff);
            border: 1px solid #dbeafe;
            box-shadow: 0 6px 20px rgba(37, 99, 235, 0.10);
        }

        .seller-hero h3 {
            margin: 0 0 0.35rem 0;
            font-size: 1.1rem;
        }

        .seller-hero p {
            margin: 0;
            color: #475569;
            font-size: 0.95rem;
        }

        .admin-card {
            border-radius: 12px;
            padding: 1rem 1.1rem;
            margin-bottom: 0.75rem;
            border: 1px solid #e5e7eb;
            background: #ffffff;
            box-shadow: 0 4px 14px rgba(15, 23, 42, 0.06);
        }

        .admin-card.pending {
            border-left: 6px solid #94a3b8;
        }

        .admin-card.approved {
            border-left: 6px solid #16a34a;
            background: #f0fdf4;
        }

        .admin-card.deleted {
            border-left: 6px solid #dc2626;
            background: #fef2f2;
        }

        .admin-title {
            font-weight: 700;
            margin-bottom: 0.3rem;
        }

        .admin-meta {
            color: #475569;
            font-size: 0.92rem;
            margin-bottom: 0.2rem;
        }

        .admin-badge {
            display: inline-block;
            margin-top: 0.35rem;
            padding: 0.2rem 0.65rem;
            border-radius: 999px;
            font-size: 0.78rem;
            font-weight: 700;
        }

        .admin-badge.pending {
            background: #e2e8f0;
            color: #1e293b;
        }

        .admin-badge.approved {
            background: #dcfce7;
            color: #166534;
        }

        .admin-badge.deleted {
            background: #fee2e2;
            color: #991b1b;
        }

        /* Nút bấm tròn, hiện đại hơn */
        .stButton>button {
            border-radius: 999px;
            padding: 0.35rem 1.1rem;
            border: none;
            background: #2563eb;
            color: white;
            font-weight: 600;
            white-space: nowrap;
            box-shadow: 0 4px 10px rgba(37, 99, 235, 0.28);
            transition: all 150ms ease;
        }

        .stButton>button:hover {
            background: #1d4ed8;
            transform: translateY(-1px);
        }

        .stButton>button:active {
            transform: translateY(0px) scale(0.99);
            box-shadow: 0 2px 6px rgba(15, 23, 42, 0.25);
        }

        .app-footer {
            margin-top: 2.2rem;
            padding-top: 1rem;
            border-top: 1px solid #e5e7eb;
            color: #64748b;
            font-size: 0.88rem;
            line-height: 1.6;
        }

        .app-footer strong {
            color: #1f2937;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_app_footer() -> None:
    st.markdown(
        """
        <div class='app-footer'>
            <div><strong>Tác giả:</strong> Nguyễn Trần Quốc Khôi, Nguyễn Trường</div>
            <div><strong>Bản quyền:</strong> © 2026 Hệ thống đăng tin & tìm kiếm nhà đất. All rights reserved.</div>
            <div>Ứng dụng phục vụ mục đích học tập, nghiên cứu và trình bày seminar.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


# =============================
# Load & lưu dữ liệu
# =============================

@st.cache_data(ttl=600)
def load_template() -> pd.DataFrame:
    return pd.read_csv(TEMPLATE_PATH)


@st.cache_data(ttl=600)
def load_data() -> pd.DataFrame:
    if os.path.exists(DATA_PATH):
        return pd.read_csv(DATA_PATH)
    return pd.DataFrame()


@st.cache_data(ttl=120)
def load_admin_posts() -> pd.DataFrame:
    if os.path.exists(ADMIN_PATH):
        df_admin = pd.read_csv(ADMIN_PATH)
        if df_admin.empty:
            return df_admin

        # Chỉ lấy bài do người bán đăng qua app (không dùng dữ liệu nền từ DATA_PATH)
        if "admin_source" not in df_admin.columns:
            return pd.DataFrame()

        return df_admin[df_admin["admin_source"].astype(str) == "seller"].reset_index(drop=True)
    return pd.DataFrame()


def save_uploaded_data(df: pd.DataFrame) -> None:
    """Lưu dữ liệu chính."""
    df.to_csv(DATA_PATH, index=False)
    load_data.clear()
    build_hybrid_matrices.clear()


def save_admin_posts(df_admin: pd.DataFrame) -> None:
    df_admin.to_csv(ADMIN_PATH, index=False)
    load_admin_posts.clear()


def append_admin_posts(df_new: pd.DataFrame) -> None:
    """Thêm bài đăng mới vào danh sách duyệt của admin, giữ toàn bộ lịch sử cũ."""
    if df_new is None or df_new.empty:
        return

    df_admin = load_admin_posts()
    if df_admin.empty:
        next_id = 1
    else:
        existing_ids = pd.to_numeric(df_admin.get("admin_post_id", pd.Series(dtype="float64")), errors="coerce")
        max_id = int(existing_ids.max()) if not existing_ids.dropna().empty else 0
        next_id = max_id + 1

    rows = df_new.copy().reset_index(drop=True)
    rows["admin_post_id"] = np.arange(next_id, next_id + len(rows))
    rows["admin_status"] = "pending"
    rows["admin_status_label"] = "Chờ duyệt"
    rows["admin_created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows["admin_source"] = "seller"

    merged = pd.concat([df_admin, rows], ignore_index=True) if not df_admin.empty else rows
    save_admin_posts(merged)


def update_admin_post_status(admin_post_id: int, status: str) -> bool:
    df_admin = load_admin_posts()
    if df_admin.empty or "admin_post_id" not in df_admin.columns:
        return False

    mask = pd.to_numeric(df_admin["admin_post_id"], errors="coerce") == admin_post_id
    if not mask.any():
        return False

    status_map = {
        "pending": "Chờ duyệt",
        "approved": "Đã duyệt",
        "deleted": "Đã xóa",
    }
    df_admin.loc[mask, "admin_status"] = status
    df_admin.loc[mask, "admin_status_label"] = status_map.get(status, "Chờ duyệt")
    df_admin.loc[mask, "admin_updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    save_admin_posts(df_admin)
    return True


def append_new_house(df_new: pd.DataFrame) -> None:
    df = load_data()
    df = pd.concat([df, df_new], ignore_index=True)
    save_uploaded_data(df)
    append_admin_posts(df_new)


def slugify_quan_name(quan_name: str) -> str:
    s = re.sub(r"[^\w\s-]", "", str(quan_name).strip().lower())
    s = re.sub(r"[\s]+", "_", s)
    return s


@st.cache_resource
def load_segment_model_bundle(quan_name: str):
    model_file = f"kmeans_k3_{slugify_quan_name(quan_name)}.joblib"
    model_path = os.path.join(SEGMENT_MODEL_DIR, model_file)
    if not os.path.exists(model_path):
        return None
    return joblib.load(model_path)


def predict_house_segment(input_data: dict[str, object]) -> tuple[str | None, str | None]:
    """Dự đoán phân khúc nhà cho dữ liệu seller nhập tay nếu đủ cột và có model theo quận."""
    quan_val = str(input_data.get("quan", "")).strip()
    if not quan_val:
        return None, "Thiếu quận nên chưa thể dự đoán phân khúc."

    bundle = load_segment_model_bundle(quan_val)
    if bundle is None:
        return None, f"Chưa tìm thấy model phân khúc cho quận '{quan_val}'."

    feature_cols = bundle.get("feature_cols", SEGMENT_FEATURE_COLS)
    row = pd.DataFrame([input_data])

    missing_cols = [c for c in feature_cols if c not in row.columns]
    if missing_cols:
        return None, f"Thiếu cột để phân cụm: {missing_cols}"

    X = row[feature_cols].apply(pd.to_numeric, errors="coerce")
    if X.isnull().any().any():
        return None, "Các cột phân cụm phải là số hợp lệ (không để trống)."

    X_scaled = bundle["scaler"].transform(X.values)
    pred_cluster = int(bundle["model"].predict(X_scaled)[0])

    cluster_label_map = bundle.get("cluster_label_map", DEFAULT_CLUSTER_LABEL_MAP)
    segment_label = (
        cluster_label_map.get(pred_cluster)
        or cluster_label_map.get(str(pred_cluster))
        or DEFAULT_CLUSTER_LABEL_MAP.get(pred_cluster, "không xác định")
    )
    return str(segment_label), None


def validate_seller_record(
    record: dict[str, object],
    available_cols: list[str],
    quan_options: list[str],
) -> list[str]:
    """Validate dữ liệu đăng tin để tránh sai kiểu/sai business rule."""
    errors: list[str] = []

    required_text_fields = ["tieu_de", "dia_chi", "quan"]
    for col in required_text_fields:
        if col in available_cols and not str(record.get(col, "")).strip():
            errors.append(f"'{col}' không được để trống.")

    if "quan" in available_cols:
        quan_val = str(record.get("quan", "")).strip()
        if quan_options and quan_val not in quan_options:
            errors.append("Giá trị 'quan' không hợp lệ.")

    positive_float_fields = ["gia_ban_num", "dien_tich_num", "chieu_ngang_num"]
    for col in positive_float_fields:
        if col in available_cols:
            val = pd.to_numeric(pd.Series([record.get(col)]), errors="coerce").iloc[0]
            if pd.isna(val) or float(val) <= 0:
                errors.append(f"'{col}' phải là số > 0.")

    non_negative_int_fields = ["so_phong_ngu", "so_phong_ve_sinh", "tong_so_tang"]
    for col in non_negative_int_fields:
        if col in available_cols:
            val = pd.to_numeric(pd.Series([record.get(col)]), errors="coerce").iloc[0]
            if pd.isna(val) or float(val) < 0 or float(val) % 1 != 0:
                errors.append(f"'{col}' phải là số nguyên không âm.")

    if "log_gia_ban" in available_cols:
        val = pd.to_numeric(pd.Series([record.get("log_gia_ban")]), errors="coerce").iloc[0]
        if pd.isna(val):
            errors.append("'log_gia_ban' phải là số hợp lệ.")

    return errors


def build_field_error_map(error_messages: list[str]) -> dict[str, str]:
    """Map lỗi dạng text thành lỗi theo field để hiển thị inline trên form."""
    field_map: dict[str, str] = {}
    for msg in error_messages:
        m = re.search(r"'([^']+)'", str(msg))
        if m:
            field = m.group(1)
            field_map[field] = str(msg)
    return field_map


def normalize_seller_record_for_import(record: dict[str, object]) -> dict[str, object]:
    """Chuẩn hóa 1 bản ghi import để hành vi gần giống form nhập tay."""
    out = dict(record)

    for text_col in ["tieu_de", "dia_chi", "mo_ta", "dac_diem", "quan", "loai_hinh", "giay_to_phap_ly", "tinh_trang_noi_that"]:
        if text_col in out:
            val = out.get(text_col)
            out[text_col] = "" if pd.isna(val) else str(val).strip()

    num_defaults = {
        "gia_ban_num": 0.0,
        "dien_tich_num": 0.0,
        "chieu_ngang_num": 0.0,
        "chieu_dai_num": 0.0,
        "log_gia_ban": np.nan,
    }
    for col, default_val in num_defaults.items():
        val = pd.to_numeric(pd.Series([out.get(col, default_val)]), errors="coerce").iloc[0]
        out[col] = float(val) if not pd.isna(val) else default_val

    # Suy diễn diện tích/chiều dài như form nhập tay
    chieu_ngang = float(out.get("chieu_ngang_num", 0) or 0)
    dien_tich = float(out.get("dien_tich_num", 0) or 0)
    chieu_dai = float(out.get("chieu_dai_num", 0) or 0)

    if chieu_ngang > 0:
        if dien_tich > 0 and chieu_dai <= 0:
            out["chieu_dai_num"] = float(dien_tich / chieu_ngang)
        elif chieu_dai > 0 and dien_tich <= 0:
            out["dien_tich_num"] = float(chieu_ngang * chieu_dai)

    # Luôn tính log_gia_ban từ gia_ban_num để đồng bộ với form web
    gia_ban_ty = float(out.get("gia_ban_num", 0) or 0)
    out["log_gia_ban"] = float(np.log1p(max(gia_ban_ty, 0.0)))

    return out


def get_segment_label_for_record(record: dict[str, object]) -> tuple[str, str | None]:
    """Trả về nhãn phân khúc cho 1 tin đăng để lưu và hiển thị phía người mua."""
    segment_label, segment_msg = predict_house_segment(record)
    if segment_label:
        return str(segment_label), None
    return "chưa xác định", segment_msg


def normalize_segment_label(value: object) -> str:
    """Chuẩn hóa nhãn phân khúc để tránh hiển thị nan/none trên UI."""
    if pd.isna(value):
        return "chưa xác định"
    text = str(value).strip()
    if text.lower() in {"", "nan", "none", "null", "<na>"}:
        return "chưa xác định"
    return text


def resolve_segment_for_display(record: dict[str, object]) -> str:
    """Lấy phân khúc để hiển thị: ưu tiên dữ liệu có sẵn, thiếu thì dự đoán KMeans từ thông tin căn nhà."""
    existing = normalize_segment_label(record.get("phan_khuc_du_doan", ""))
    if existing != "chưa xác định":
        return existing

    predicted, _ = get_segment_label_for_record(record)
    return normalize_segment_label(predicted)


def is_effectively_blank_import_row(
    record: dict[str, object],
    ignore_cols: set[str] | None = None,
) -> bool:
    """True nếu dòng trống ở mọi cột (trừ các cột được bỏ qua)."""
    ignored = ignore_cols or set()
    for key, value in record.items():
        if str(key) in ignored:
            continue
        if pd.isna(value):
            continue
        if isinstance(value, str):
            if value.strip() == "":
                continue
            return False
        return False
    return True


def build_excel_import_columns(base_cols: list[str]) -> list[str]:
    """Đảm bảo template/import Excel luôn có đủ cột bắt buộc cho nghiệp vụ."""
    cols = [str(c).strip() for c in base_cols if str(c).strip()]
    for c in EXCEL_REQUIRED_EXTRA_COLS:
        if c not in cols:
            cols.append(c)
    return cols


def build_seller_excel_template_bytes(
    columns: list[str],
    quan_options: list[str],
    loai_hinh_options: list[str],
    phap_ly_options: list[str],
    noi_that_options: list[str],
    max_rows: int = 500,
) -> tuple[bytes | None, str | None]:
    """Tạo file Excel mẫu nhập liệu có dropdown + ràng buộc kiểu dữ liệu."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except Exception:
        return None, "Thiếu thư viện openpyxl để tạo file Excel mẫu."

    wb = Workbook()
    ws = wb.active
    ws.title = "Nhap_lieu"

    header_fill = PatternFill("solid", fgColor="E8F0FE")
    for cidx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=cidx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(cidx)].width = max(14, min(35, len(str(col_name)) + 4))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    guide = wb.create_sheet("Huong_dan")
    guide["A1"] = "HƯỚNG DẪN NHẬP LIỆU"
    guide["A1"].font = Font(bold=True)
    guide["A3"] = "1) Nhập dữ liệu tại sheet 'Nhap_lieu'."
    guide["A4"] = "2) Các cột Quận/Loại hình/Giấy tờ pháp lý/Nội thất dùng dropdown như form web."
    guide["A5"] = "3) Các cột số phải >= 0, riêng gia_ban_num/dien_tich_num/chieu_ngang_num phải > 0."
    guide["A6"] = "4) Có thể nhập dien_tich_num hoặc chieu_dai_num; hệ thống sẽ suy diễn giá trị còn lại nếu chieu_ngang_num > 0."
    guide["A7"] = "5) Không đổi tên cột ở dòng tiêu đề."
    guide.column_dimensions["A"].width = 120

    enum_sheet = wb.create_sheet("Danh_muc")
    enum_sheet.sheet_state = "hidden"
    enum_map = {
        "quan": sorted([str(x).strip() for x in quan_options if str(x).strip()]),
        "loai_hinh": sorted([str(x).strip() for x in loai_hinh_options if str(x).strip()]),
        "giay_to_phap_ly": sorted([str(x).strip() for x in phap_ly_options if str(x).strip()]),
        "tinh_trang_noi_that": sorted([str(x).strip() for x in noi_that_options if str(x).strip()]),
    }

    enum_col_idx: dict[str, int] = {}
    cur_col = 1
    for key, values in enum_map.items():
        if not values:
            continue
        enum_col_idx[key] = cur_col
        enum_sheet.cell(row=1, column=cur_col, value=key)
        for ridx, v in enumerate(values, start=2):
            enum_sheet.cell(row=ridx, column=cur_col, value=v)
        cur_col += 1

    col_to_idx = {name: idx + 1 for idx, name in enumerate(columns)}
    row_start = 2
    row_end = max_rows + 1

    for field in ["quan", "loai_hinh", "giay_to_phap_ly", "tinh_trang_noi_that"]:
        if field in col_to_idx and field in enum_col_idx and enum_map[field]:
            enum_col_letter = get_column_letter(enum_col_idx[field])
            enum_last_row = len(enum_map[field]) + 1
            formula = f"=Danh_muc!${enum_col_letter}$2:${enum_col_letter}${enum_last_row}"
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            ws.add_data_validation(dv)
            field_col = get_column_letter(col_to_idx[field])
            dv.add(f"{field_col}{row_start}:{field_col}{row_end}")

    positive_float_fields = ["gia_ban_num", "dien_tich_num", "chieu_ngang_num"]
    non_negative_fields = [
        "chieu_dai_num",
        "dien_tich_dat_num",
        "dien_tich_su_dung_num",
        "gia_m2_num",
        "so_phong_ngu",
        "so_phong_ve_sinh",
        "tong_so_tang",
        "log_gia_ban",
    ]

    for field in positive_float_fields:
        if field in col_to_idx:
            col_letter = get_column_letter(col_to_idx[field])
            dv = DataValidation(type="decimal", operator="greaterThan", formula1="0", allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}{row_start}:{col_letter}{row_end}")

    for field in non_negative_fields:
        if field in col_to_idx:
            col_letter = get_column_letter(col_to_idx[field])
            dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}{row_start}:{col_letter}{row_end}")

    # Cột log_gia_ban tự lấy log(1 + gia_ban_num) trực tiếp trong Excel
    if "gia_ban_num" in col_to_idx and "log_gia_ban" in col_to_idx:
        gia_col = get_column_letter(col_to_idx["gia_ban_num"])
        log_col = get_column_letter(col_to_idx["log_gia_ban"])
        for r in range(row_start, row_end + 1):
            ws[f"{log_col}{r}"] = f"=IFERROR(LN(1+{gia_col}{r}),\"\")"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), None


# =============================
# Hàm hỗ trợ recommendation
# =============================

def clean_text(s: str) -> str:
    """Làm sạch text cho content-based recommendation."""
    s = html.unescape(str(s or ""))
    s = re.sub(r"<[^>]+>", " ", s)
    s = s.lower()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


@st.cache_data(ttl=3600)
def build_tfidf_full(df_hash: int):
    """Xây TF-IDF cho toàn bộ catalog, dùng cho các tác vụ offline nếu cần.
    (Hiện tại không dùng trực tiếp trong UI nhưng giữ lại để mở rộng sau.)
    """
    df = load_data()
    if df.empty:
        return None, None

    text_cols = [
        "tieu_de",
        "dia_chi",
        "mo_ta",
        "dac_diem",
        "quan",
        "loai_hinh",
        "giay_to_phap_ly",
        "tinh_trang_noi_that",
    ]

    combined_text = []
    for _, row in df.iterrows():
        parts = [str(row.get(c, "")) for c in text_cols if c in df.columns]
        combined_text.append(clean_text(" ".join(parts)))

    if not any(combined_text):
        return None, None

    tfidf = TfidfVectorizer(
        min_df=2,
        max_df=0.95,
        ngram_range=(1, 2),
        max_features=10000,
    )
    X = tfidf.fit_transform(combined_text)
    return tfidf, X


@st.cache_data(ttl=3600)
def build_hybrid_matrices():
    """Tiền tính các ma trận similarity cho Hybrid Recommendation."""
    df = load_data()
    if df.empty:
        return None, None, None

    # 1) Content similarity (TF‑IDF + cosine)
    text_cols = [
        "tieu_de",
        "dia_chi",
        "mo_ta",
        "dac_diem",
        "quan",
        "loai_hinh",
        "giay_to_phap_ly",
        "tinh_trang_noi_that",
    ]

    combined_text = []
    for _, row in df.iterrows():
        parts = [str(row.get(c, "")) for c in text_cols if c in df.columns]
        combined_text.append(clean_text(" ".join(parts)))

    if not any(combined_text):
        return None, None, None

    tfidf = TfidfVectorizer(
        min_df=2,
        max_df=0.95,
        ngram_range=(1, 2),
        max_features=10000,
    )
    X_text = tfidf.fit_transform(combined_text)
    content_sim = cosine_similarity(X_text, dense_output=False).toarray().astype(float)

    # 2) Price similarity
    price = pd.to_numeric(df.get("gia_ban_num", 0), errors="coerce").fillna(0).values.astype(float)
    price_scaled = (price - price.min()) / (price.max() - price.min() + 1e-9)
    price_dist = np.abs(price_scaled[:, None] - price_scaled[None, :])
    price_sim = 1.0 / (1.0 + price_dist)

    # 3) Location similarity (quận)
    quan_ohe = pd.get_dummies(df.get("quan", "unknown").fillna("unknown").astype(str), dtype=float).values
    location_dist = np.linalg.norm(quan_ohe[:, None, :] - quan_ohe[None, :, :], axis=2)
    location_sim = 1.0 / (1.0 + location_dist)

    def minmax_norm(m: np.ndarray) -> np.ndarray:
        row_min = m.min(axis=1, keepdims=True)
        row_max = m.max(axis=1, keepdims=True)
        return (m - row_min) / (row_max - row_min + 1e-9)

    content_sim_n = minmax_norm(content_sim)
    price_sim_n = minmax_norm(price_sim)
    location_sim_n = minmax_norm(location_sim)

    return content_sim_n, price_sim_n, location_sim_n


def recommend_content_based(query_text: str, df: pd.DataFrame, top_k: int = 5) -> pd.DataFrame:
    """Đề xuất theo nội dung (content-based) trên dataframe đã lọc df."""
    if df is None or df.empty:
        return pd.DataFrame()

    text_cols = [
        "tieu_de",
        "dia_chi",
        "mo_ta",
        "dac_diem",
        "quan",
        "loai_hinh",
        "giay_to_phap_ly",
        "tinh_trang_noi_that",
    ]

    combined_text: list[str] = []
    for _, row in df.iterrows():
        parts = [str(row.get(c, "")) for c in text_cols if c in df.columns]
        combined_text.append(clean_text(" ".join(parts)))

    if not any(combined_text):
        return pd.DataFrame()

    try:
        tfidf = TfidfVectorizer(
            min_df=1,
            max_df=1.0,
            ngram_range=(1, 2),
            max_features=5000,
        )
        X_tfidf = tfidf.fit_transform(combined_text)
    except Exception as exc:  # phòng lỗi hiếm
        st.warning(f"Lỗi khi tạo TF-IDF: {exc}")
        return pd.DataFrame()

    query_clean = clean_text(query_text)
    if not query_clean:
        return pd.DataFrame()

    try:
        q_vec = tfidf.transform([query_clean])
    except Exception as exc:
        st.warning(f"Lỗi khi transform query: {exc}")
        return pd.DataFrame()

    sim = cosine_similarity(q_vec, X_tfidf)[0]
    if sim.size == 0:
        return pd.DataFrame()

    k = min(top_k, len(df))
    top_idx = np.argsort(sim)[::-1][:k]

    df_reset = df.reset_index(drop=True)
    out = df_reset.iloc[top_idx].copy()
    out["score_content"] = sim[top_idx]
    return out.reset_index(drop=True)


def recommend_hybrid(
    item_idx: int,
    df: pd.DataFrame,
    top_k: int = 5,
    w_content: float = 0.5,
    w_price: float = 0.3,
    w_location: float = 0.2,
) -> pd.DataFrame:
    """Đề xuất Hybrid (Content + Price + Location) cho 1 item trong df."""
    if df is None or df.empty or item_idx < 0 or item_idx >= len(df):
        return pd.DataFrame()

    # Ma trận hybrid được cache theo toàn bộ dataset, tránh tính toán lại mỗi lần xem chi tiết
    matrices = build_hybrid_matrices()
    if any(m is None for m in matrices):
        return pd.DataFrame()

    content_sim_n, price_sim_n, location_sim_n = matrices

    scores = (
        w_content * content_sim_n[item_idx]
        + w_price * price_sim_n[item_idx]
        + w_location * location_sim_n[item_idx]
    )
    scores[item_idx] = -1.0

    k = min(top_k, len(df))
    top_idx = np.argsort(scores)[::-1][:k]

    out = df.iloc[top_idx].copy()
    out["score_hybrid"] = scores[top_idx]
    out["score_content"] = content_sim_n[item_idx, top_idx]
    out["score_price"] = price_sim_n[item_idx, top_idx]
    out["score_location"] = location_sim_n[item_idx, top_idx]

    return out.reset_index(drop=True)


# =============================
# Giao diện người bán
# =============================


def seller_interface() -> None:
    st.header("📝 Đăng thông tin ngôi nhà")
    st.markdown(
        """
        <div class='seller-hero'>
            <h3>Đăng tin nhanh chóng & chuyên nghiệp</h3>
            <p>Bạn có thể nhập trực tiếp từng tin hoặc import hàng loạt bằng file Excel mẫu.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    tab1, tab2 = st.tabs(["✍️ Nhập trực tiếp", "📥 Import từ file"])

    with tab1:
        template = load_template()
        cols = template.columns.tolist()
        excel_cols = build_excel_import_columns(cols)
        base_df = load_data()
        field_errors: dict[str, str] = st.session_state.get("seller_field_errors", {})

        def get_options(col_name: str) -> list[str]:
            if base_df.empty or col_name not in base_df.columns:
                return []
            vals = (
                base_df[col_name]
                .dropna()
                .astype(str)
                .str.strip()
            )
            vals = vals[vals != ""]
            return sorted(vals.unique().tolist())

        quan_options = get_options("quan")
        if not quan_options:
            quan_options = ["Binh Thanh", "Go Vap", "Phu Nhuan"]

        loai_hinh_options = get_options("loai_hinh")
        phap_ly_options = get_options("giay_to_phap_ly")
        noi_that_options = get_options("tinh_trang_noi_that")

        numeric_field_candidates = {
            "gia_ban_num",
            "dien_tich_num",
            "so_phong_ngu",
            "so_phong_ve_sinh",
            "tong_so_tang",
            "chieu_ngang_num",
            "chieu_dai_num",
            "dien_tich_dat_num",
            "dien_tich_su_dung_num",
            "gia_m2_num",
            "log_gia_ban",
        }

        vn_labels = {
            "tieu_de": "Tiêu đề",
            "dia_chi": "Địa chỉ",
            "mo_ta": "Mô tả",
            "dac_diem": "Đặc điểm",
            "quan": "Quận",
            "loai_hinh": "Loại hình",
            "giay_to_phap_ly": "Giấy tờ pháp lý",
            "tinh_trang_noi_that": "Tình trạng nội thất",
            "dien_tich_num": "Diện tích (m²)",
            "gia_ban_num": "Giá bán (tỷ VNĐ)",
            "chieu_ngang_num": "Chiều ngang (m)",
            "chieu_dai_num": "Chiều dài (m)",
            "log_gia_ban": "Log giá bán (tự tính)",
        }

        with st.form("seller_direct_form", clear_on_submit=False):
            st.markdown("<div class='section-title'>📌 Thông tin chi tiết căn nhà</div>", unsafe_allow_html=True)

            col_left, col_right = st.columns(2)
            input_data: dict[str, object] = {}

            for idx, col in enumerate(cols):
                target_col = col_left if idx % 2 == 0 else col_right
                label = vn_labels.get(col, col.replace("_", " ").title())

                with target_col:
                    if col == "quan":
                        input_data[col] = st.selectbox(
                            "Quận",
                            options=quan_options,
                            key=f"seller_form_{col}",
                        )
                        if col in field_errors:
                            st.markdown(f"<span style='color:#dc2626'>{field_errors[col]}</span>", unsafe_allow_html=True)
                    elif col == "loai_hinh" and loai_hinh_options:
                        input_data[col] = st.selectbox(
                            label,
                            options=loai_hinh_options,
                            key=f"seller_form_{col}",
                        )
                        if col in field_errors:
                            st.markdown(f"<span style='color:#dc2626'>{field_errors[col]}</span>", unsafe_allow_html=True)
                    elif col == "giay_to_phap_ly" and phap_ly_options:
                        input_data[col] = st.selectbox(
                            label,
                            options=phap_ly_options,
                            key=f"seller_form_{col}",
                        )
                        if col in field_errors:
                            st.markdown(f"<span style='color:#dc2626'>{field_errors[col]}</span>", unsafe_allow_html=True)
                    elif col == "tinh_trang_noi_that" and noi_that_options:
                        input_data[col] = st.selectbox(
                            label,
                            options=noi_that_options,
                            key=f"seller_form_{col}",
                        )
                        if col in field_errors:
                            st.markdown(f"<span style='color:#dc2626'>{field_errors[col]}</span>", unsafe_allow_html=True)
                    elif col in numeric_field_candidates:
                        input_data[col] = st.number_input(
                            label,
                            min_value=0.0,
                            value=0.0,
                            step=1.0,
                            format="%.4f",
                            key=f"seller_form_{col}",
                        )
                        if col in field_errors:
                            st.markdown(f"<span style='color:#dc2626'>{field_errors[col]}</span>", unsafe_allow_html=True)
                    elif col in {"mo_ta", "dac_diem"}:
                        input_data[col] = st.text_area(label, height=110, key=f"seller_form_{col}")
                        if col in field_errors:
                            st.markdown(f"<span style='color:#dc2626'>{field_errors[col]}</span>", unsafe_allow_html=True)
                    else:
                        input_data[col] = st.text_input(label, key=f"seller_form_{col}")
                        if col in field_errors:
                            st.markdown(f"<span style='color:#dc2626'>{field_errors[col]}</span>", unsafe_allow_html=True)

            # Bổ sung input bắt buộc cho phân cụm nếu template chưa có
            extra_numeric_map = {
                "chieu_ngang_num": "Chiều ngang (m)",
                "chieu_dai_num": "Chiều dài (m)",
            }
            for extra_col, extra_label in extra_numeric_map.items():
                if extra_col not in input_data:
                    input_data[extra_col] = st.number_input(
                        extra_label,
                        min_value=0.0,
                        value=0.0,
                        step=0.1,
                        format="%.4f",
                        key=f"seller_extra_{extra_col}",
                    )
                    if extra_col in field_errors:
                        st.markdown(f"<span style='color:#dc2626'>{field_errors[extra_col]}</span>", unsafe_allow_html=True)

            st.caption("💡 Giá bán nhập theo đơn vị **tỷ VNĐ**. Hệ thống sẽ tự tính `log_gia_ban`. Nếu chỉ nhập diện tích hoặc chiều dài, hệ thống sẽ tự suy ra giá trị còn lại dựa trên chiều ngang.")

            submitted = st.form_submit_button("🚀 Đăng nhà")

        if submitted:
            st.session_state["seller_field_errors"] = {}

            if not str(input_data.get("tieu_de", "")).strip():
                st.session_state["seller_field_errors"] = {"tieu_de": "'tieu_de' không được để trống."}
                st.warning("Vui lòng nhập tiêu đề trước khi đăng.")
                st.rerun()
            else:
                # Kiểm tra kiểu dữ liệu số cho các cột số trước khi lưu
                numeric_cols_in_form = [c for c in cols if c in numeric_field_candidates]
                bad_numeric_cols = []
                for nc in numeric_cols_in_form:
                    val = pd.to_numeric(pd.Series([input_data.get(nc)]), errors="coerce").iloc[0]
                    if pd.isna(val):
                        bad_numeric_cols.append(nc)
                    else:
                        input_data[nc] = float(val)

                # Chuẩn hóa và suy diễn các cột hình học
                chieu_ngang = float(pd.to_numeric(pd.Series([input_data.get("chieu_ngang_num", 0)]), errors="coerce").iloc[0] or 0)
                dien_tich = float(pd.to_numeric(pd.Series([input_data.get("dien_tich_num", 0)]), errors="coerce").iloc[0] or 0)
                chieu_dai = float(pd.to_numeric(pd.Series([input_data.get("chieu_dai_num", 0)]), errors="coerce").iloc[0] or 0)

                if chieu_ngang > 0:
                    if dien_tich > 0 and chieu_dai <= 0:
                        chieu_dai = dien_tich / chieu_ngang
                        input_data["chieu_dai_num"] = float(chieu_dai)
                    elif chieu_dai > 0 and dien_tich <= 0:
                        dien_tich = chieu_ngang * chieu_dai
                        input_data["dien_tich_num"] = float(dien_tich)

                # Tính log_gia_ban từ giá (đơn vị tỷ VNĐ)
                gia_ban_ty = float(pd.to_numeric(pd.Series([input_data.get("gia_ban_num", 0)]), errors="coerce").iloc[0] or 0)
                if gia_ban_ty > 0:
                    input_data["log_gia_ban"] = float(np.log1p(gia_ban_ty))

                if bad_numeric_cols:
                    st.session_state["seller_field_errors"] = {
                        c: f"'{c}' không hợp lệ, vui lòng nhập số đúng định dạng."
                        for c in bad_numeric_cols
                    }
                    st.error(f"Các cột số không hợp lệ: {bad_numeric_cols}. Vui lòng nhập lại đúng kiểu số.")
                    st.rerun()

                if input_data.get("chieu_ngang_num", 0) <= 0:
                    st.session_state["seller_field_errors"] = {
                        "chieu_ngang_num": "'chieu_ngang_num' phải > 0."
                    }
                    st.error("'chieu_ngang_num' phải > 0 để suy ra diện tích/chiều dài và phân cụm.")
                    st.rerun()

                if float(input_data.get("dien_tich_num", 0) or 0) <= 0:
                    st.session_state["seller_field_errors"] = {
                        "dien_tich_num": "'dien_tich_num' cần > 0 (hoặc nhập chiều dài để hệ thống tự suy ra).",
                        "chieu_dai_num": "'chieu_dai_num' cần > 0 nếu chưa nhập diện tích.",
                    }
                    st.error("Bạn cần nhập ít nhất một trong hai: diện tích hoặc chiều dài (kèm chiều ngang > 0).")
                    st.rerun()

                record_errors = validate_seller_record(
                    input_data,
                    available_cols=list(set(cols + ["chieu_ngang_num", "chieu_dai_num", "dien_tich_num", "log_gia_ban"])),
                    quan_options=quan_options,
                )
                if record_errors:
                    st.session_state["seller_field_errors"] = build_field_error_map(record_errors)
                    st.error("Không thể đăng tin do dữ liệu chưa hợp lệ:")
                    for e in record_errors:
                        st.write(f"- {e}")
                    st.rerun()

                segment_label, segment_msg = get_segment_label_for_record(input_data)
                input_data["phan_khuc_du_doan"] = segment_label

                df_new = pd.DataFrame([input_data])
                append_new_house(df_new)

                df_show = df_new.copy()
                if segment_label and segment_label != "chưa xác định":
                    st.success(f"🏷️ Phân khúc dự đoán: **{segment_label}**")
                elif segment_msg:
                    st.info(f"ℹ️ {segment_msg}")

                st.session_state["latest_df_new"] = df_show
                st.session_state["seller_field_errors"] = {}
                st.success("✅ Đăng thông tin thành công!")
                st.markdown("<div class='section-title'>Thông tin vừa đăng</div>", unsafe_allow_html=True)
                st.dataframe(df_show, use_container_width=True)

    with tab2:
        st.markdown("<div class='section-title'>📁 Import dữ liệu hàng loạt</div>", unsafe_allow_html=True)

        excel_template_bytes, template_err = build_seller_excel_template_bytes(
            columns=excel_cols,
            quan_options=quan_options,
            loai_hinh_options=loai_hinh_options,
            phap_ly_options=phap_ly_options,
            noi_that_options=noi_that_options,
            max_rows=500,
        )
        if template_err:
            st.error(template_err)
        else:
            st.download_button(
                "Tải file mẫu Excel",
                excel_template_bytes,
                file_name="house_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        uploaded_file = st.file_uploader("Chọn file Excel để import", type=["xlsx", "xls"])
        if uploaded_file is not None:
            try:
                df_new = pd.read_excel(uploaded_file)
            except Exception as exc:
                st.error(f"Không thể đọc file Excel: {exc}")
                return

            if df_new.empty:
                st.warning("File Excel chưa có dữ liệu để import.")
                return

            df_new.columns = [str(c).strip() for c in df_new.columns]
            missing_required_cols = [c for c in excel_cols if c not in df_new.columns]
            if missing_required_cols:
                st.error("File Excel thiếu cột bắt buộc, hệ thống chưa lưu.")
                st.write(f"Thiếu các cột: {missing_required_cols}")
                return

            df_new = df_new[excel_cols].copy()

            # Validate toàn bộ dòng import, nếu có lỗi thì chặn lưu
            import_errors: list[str] = []
            cleaned_rows: list[dict[str, object]] = []
            skipped_blank_rows = 0
            for ridx, row in df_new.iterrows():
                raw_row_dict = row.to_dict()
                if is_effectively_blank_import_row(raw_row_dict, ignore_cols={"log_gia_ban"}):
                    skipped_blank_rows += 1
                    continue

                row_dict = normalize_seller_record_for_import(raw_row_dict)
                errs = validate_seller_record(
                    row_dict,
                    available_cols=list(set(df_new.columns.tolist() + ["chieu_dai_num", "chieu_ngang_num", "dien_tich_num", "log_gia_ban"])),
                    quan_options=quan_options,
                )
                if errs:
                    import_errors.append(f"Dòng {ridx + 1}: " + "; ".join(errs))
                else:
                    segment_label, _ = get_segment_label_for_record(row_dict)
                    row_dict["phan_khuc_du_doan"] = segment_label
                    cleaned_rows.append(row_dict)

            if import_errors:
                st.error("File import có dữ liệu không hợp lệ, hệ thống chưa lưu.")
                st.write("Chi tiết lỗi (tối đa 20 dòng):")
                for msg in import_errors[:20]:
                    st.write(f"- {msg}")
                if len(import_errors) > 20:
                    st.write(f"... và {len(import_errors) - 20} lỗi khác")
                return

            if not cleaned_rows:
                if skipped_blank_rows > 0:
                    st.info("Không có dữ liệu để import. Các dòng trống đã được tự động bỏ qua.")
                else:
                    st.warning("Không có dòng hợp lệ để import.")
                return

            df_new = pd.DataFrame(cleaned_rows)

            append_new_house(df_new)
            st.session_state["latest_df_new"] = df_new
            st.success(f"✅ Import thành công {len(df_new)} căn nhà!")
            if skipped_blank_rows > 0:
                st.caption(f"Đã bỏ qua {skipped_blank_rows} dòng trống.")
            st.markdown("<div class='section-title'>Dữ liệu vừa import</div>", unsafe_allow_html=True)
            st.dataframe(df_new, use_container_width=True)


# =============================
# Giao diện người mua
# =============================


def buyer_interface() -> None:
    st.header("🏡 Tìm kiếm và xem thông tin nhà")
    base_df = load_data()
    if base_df.empty:
        st.warning("Chưa có dữ liệu nhà nào.")
        return

    st.markdown("<div class='section-title'>🔎 Tìm kiếm nâng cao</div>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)

    with col1:
        quan = st.selectbox(
            "Quận", [""] + sorted(base_df.get("quan", []).dropna().unique().tolist())
        )

    with col2:
        min_area = st.number_input("Diện tích tối thiểu", min_value=0.0, value=0.0)
        max_area = st.number_input("Diện tích tối đa", min_value=0.0, value=0.0)

    with col3:
        min_price = st.number_input("Giá bán tối thiểu", min_value=0.0, value=0.0)
        max_price = st.number_input("Giá bán tối đa", min_value=0.0, value=0.0)

    text_query = st.text_input("Tìm kiếm theo nội dung mô tả")

    # Lọc dữ liệu
    df = base_df.copy()
    if quan:
        df = df[df["quan"] == quan]
    if min_area > 0:
        df = df[df["dien_tich_num"] >= min_area]
    if max_area > 0:
        df = df[df["dien_tich_num"] <= max_area]
    if min_price > 0:
        df = df[df["gia_ban_num"] >= min_price]
    if max_price > 0:
        df = df[df["gia_ban_num"] <= max_price]

    if df.empty:
        st.warning("Không có căn nhà nào phù hợp với bộ lọc hiện tại.")
        return

    # Khi mới vào (không lọc/không tìm kiếm), trộn ngẫu nhiên danh sách để hiển thị đa dạng
    is_default_view = (
        not quan
        and min_area <= 0
        and max_area <= 0
        and min_price <= 0
        and max_price <= 0
        and not text_query.strip()
    )
    if is_default_view:
        if "buyer_random_seed" not in st.session_state:
            st.session_state["buyer_random_seed"] = int(np.random.randint(0, 1_000_000_000))
        df = df.sample(frac=1.0, random_state=st.session_state["buyer_random_seed"]).reset_index(drop=True)
    else:
        df = df.reset_index(drop=True)

    if text_query:
        st.info("🔍 Đang sử dụng hệ thống đề xuất theo nội dung (Content-based)")
        with st.spinner("Đang tìm kiếm căn nhà phù hợp..."):
            rec_df = recommend_content_based(text_query, df, top_k=len(df))

        if rec_df.empty:
            st.warning("Không tìm thấy căn nhà phù hợp với nội dung mô tả.")
            return

        st.subheader("Kết quả đề xuất theo nội dung")
        show_cols = [
            c
            for c in [
                "tieu_de",
                "quan",
                "loai_hinh",
                "gia_ban_num",
                "dien_tich_num",
                "score_content",
            ]
            if c in rec_df.columns
        ]
        st.dataframe(rec_df[show_cols].head(LIST_PAGE_SIZE), use_container_width=True)

        # Dùng danh sách đề xuất làm danh sách hiển thị
        df = rec_df.drop(columns=["score_content"], errors="ignore")

    st.markdown("<div class='section-title'>📋 Danh sách nhà</div>", unsafe_allow_html=True)

    criteria_signature = (
        quan,
        float(min_area),
        float(max_area),
        float(min_price),
        float(max_price),
        text_query.strip().lower(),
    )

    if st.session_state.get("buyer_criteria_signature") != criteria_signature:
        st.session_state["buyer_criteria_signature"] = criteria_signature
        st.session_state["buyer_visible_count"] = LIST_PAGE_SIZE

    visible_count = int(st.session_state.get("buyer_visible_count", LIST_PAGE_SIZE))
    show_count = min(visible_count, len(df))
    df_show = df.head(show_count).reset_index(drop=True)
    st.caption(f"Đang hiển thị {show_count}/{len(df)} căn nhà")

    for idx, row in df_show.iterrows():
        title = str(row.get("tieu_de", ""))
        quan_val = str(row.get("quan", ""))
        loai_hinh = str(row.get("loai_hinh", ""))
        dien_tich = row.get("dien_tich_num", "")
        gia_ban = row.get("gia_ban_num", "")
        row_dict = row.to_dict()
        segment_text = resolve_segment_for_display(row_dict)
        row_dict["phan_khuc_du_doan"] = segment_text

        with st.container(border=True):
            top_left, top_right = st.columns([5, 1])
            with top_left:
                st.markdown(f"<div class='house-card-title'>{title}</div>", unsafe_allow_html=True)
            with top_right:
                st.markdown("<div style='height:0.15rem'></div>", unsafe_allow_html=True)
                if st.button("Xem chi tiết", key=f"house_btn_{idx}"):
                    # Tìm index tương ứng trong toàn bộ dataframe để dùng cho hybrid
                    mask = (
                        base_df["tieu_de"].fillna("").astype(str).str.strip()
                        == str(row.get("tieu_de", "")).strip()
                    ) & (
                        base_df["dia_chi"].fillna("").astype(str).str.strip()
                        == str(row.get("dia_chi", "")).strip()
                    )
                    match_idx = base_df.index[mask].tolist()

                    st.session_state["selected_house"] = row_dict
                    st.session_state["selected_house_idx"] = match_idx[0] if match_idx else None
                    st.session_state["page"] = "house_details"
                    st.rerun()

            st.markdown(
                f"""
                <div class='house-card-meta'>📍 {quan_val} • 🏷 {loai_hinh} • 📐 {dien_tich} m²</div>
                <div class='house-card-meta'>🧩 Phân khúc: <b>{segment_text}</b></div>
                <div class='house-price'>💰 {gia_ban} tỷ VNĐ</div>
                """,
                unsafe_allow_html=True,
            )

    if show_count < len(df):
        if st.button("Xem thêm", key="buyer_show_more"):
            st.session_state["buyer_visible_count"] = min(show_count + LIST_PAGE_SIZE, len(df))
            st.rerun()


# =============================
# Giao diện chi tiết ngôi nhà
# =============================


def house_details_interface() -> None:
    st.header("📄 Chi tiết ngôi nhà")

    if "selected_house" not in st.session_state or not st.session_state["selected_house"]:
        st.warning("Không có thông tin ngôi nhà nào được chọn.")
        if st.button("Quay lại", key="back_empty"):
            st.session_state["page"] = "buyer"
            st.rerun()
        return

    house = st.session_state["selected_house"]
    house["phan_khuc_du_doan"] = resolve_segment_for_display(house)

    left_col, right_col = st.columns([1.05, 1], gap="large")

    with left_col:
        st.markdown("<div class='section-title'>📌 Thông tin chi tiết</div>", unsafe_allow_html=True)
        info_left, info_right = st.columns(2)
        with info_left:
            st.write(f"**Tiêu đề:** {house.get('tieu_de', '')}")
            st.write(f"**Quận:** {house.get('quan', '')}")
            st.write(f"**Loại hình:** {house.get('loai_hinh', '')}")
            st.write(f"**Diện tích:** {house.get('dien_tich_num', '')} m²")
            st.write(f"**Phân khúc:** {normalize_segment_label(house.get('phan_khuc_du_doan', ''))}")
        with info_right:
            st.write(f"**Giá bán:** {house.get('gia_ban_num', '')} tỷ VNĐ")
            st.write(f"**Địa chỉ:** {house.get('dia_chi', '')}")
            st.write(f"**Giấy tờ pháp lý:** {house.get('giay_to_phap_ly', '')}")
            st.write(f"**Tình trạng nội thất:** {house.get('tinh_trang_noi_that', '')}")

        st.write(f"**Mô tả:** {house.get('mo_ta', '')}")
        st.write(f"**Đặc điểm:** {house.get('dac_diem', '')}")

        st.markdown("---")
        if st.button("Quay lại", key="back_button"):
            st.session_state["selected_house"] = None
            st.session_state["selected_house_idx"] = None
            st.session_state["page"] = "buyer"
            st.rerun()

    with right_col:
        st.markdown("<div class='section-title'>🤝 Căn nhà tương tự (Đề xuất Hybrid)</div>", unsafe_allow_html=True)

        df = load_data()
        item_idx = st.session_state.get("selected_house_idx")

        if df.empty or item_idx is None:
            st.info("Không đủ thông tin để tính toán đề xuất.")
        else:
            with st.spinner("Đang tìm kiếm căn nhà tương tự..."):
                rec_df = recommend_hybrid(item_idx, df, top_k=5)

            if rec_df.empty:
                st.info("Không tìm thấy căn nhà tương tự.")
            else:
                rec_scroll_box = st.container(height=650, border=False)
                with rec_scroll_box:
                    for ridx, r in rec_df.iterrows():
                        with st.container(border=True):
                            st.markdown(f"**{r.get('tieu_de', '')}**")
                            st.caption(
                                f"📍 {r.get('quan', '')} • 🏷 {r.get('loai_hinh', '')} • 📐 {r.get('dien_tich_num', '')} m²"
                            )
                            st.markdown(f"<div class='house-price'>💰 {r.get('gia_ban_num', '')} tỷ VNĐ</div>", unsafe_allow_html=True)

                            action_spacer, action_col = st.columns([6, 2])
                            with action_spacer:
                                st.empty()
                            with action_col:
                                if st.button("Chi tiết", key=f"rec_detail_{ridx}", type="secondary"):
                                    # Cập nhật nhà được chọn và index rồi render lại trang chi tiết
                                    mask = (
                                        df["tieu_de"].fillna("").astype(str).str.strip()
                                        == str(r.get("tieu_de", "")).strip()
                                    ) & (
                                        df["dia_chi"].fillna("").astype(str).str.strip()
                                        == str(r.get("dia_chi", "")).strip()
                                    )
                                    match_idx = df.index[mask].tolist()

                                    st.session_state["selected_house"] = r.to_dict()
                                    st.session_state["selected_house_idx"] = match_idx[0] if match_idx else None
                                    st.rerun()


# =============================
# Giao diện Admin
# =============================


def admin_interface() -> None:
    st.header("🛡️ Quản lý tin đăng (Admin)")
    st.markdown(
        """
        <div class='seller-hero'>
            <h3>Kiểm duyệt tin đăng chuyên nghiệp</h3>
            <p>Duyệt hoặc xóa từng bài đăng. Trạng thái sẽ được lưu lại và hiển thị màu trực quan.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    df_admin = load_admin_posts()
    if df_admin.empty:
        st.info("Chưa có bài đăng nào trong danh sách kiểm duyệt.")
        return

    # Đảm bảo các cột trạng thái luôn tồn tại cho dữ liệu cũ
    if "admin_status" not in df_admin.columns:
        df_admin["admin_status"] = "pending"
    if "admin_status_label" not in df_admin.columns:
        df_admin["admin_status_label"] = "Chờ duyệt"
    if "admin_post_id" not in df_admin.columns:
        df_admin["admin_post_id"] = np.arange(1, len(df_admin) + 1)
        save_admin_posts(df_admin)

    pending_count = int((df_admin["admin_status"] == "pending").sum())
    approved_count = int((df_admin["admin_status"] == "approved").sum())
    deleted_count = int((df_admin["admin_status"] == "deleted").sum())

    m1, m2, m3 = st.columns(3)
    m1.metric("Chờ duyệt", pending_count)
    m2.metric("Đã duyệt", approved_count)
    m3.metric("Đã xóa", deleted_count)

    st.markdown("<div class='section-title'>📚 Danh sách bài đăng</div>", unsafe_allow_html=True)

    display_df = df_admin.sort_values(by="admin_post_id", ascending=False).reset_index(drop=True)
    status_text_map = {
        "pending": "Chờ duyệt",
        "approved": "Đã duyệt",
        "deleted": "Đã xóa",
    }

    for _, row in display_df.iterrows():
        admin_post_id = int(row.get("admin_post_id", 0))
        status_key = str(row.get("admin_status", "pending"))
        if status_key not in {"pending", "approved", "deleted"}:
            status_key = "pending"

        title = str(row.get("tieu_de", "Không có tiêu đề"))
        quan = str(row.get("quan", ""))
        loai_hinh = str(row.get("loai_hinh", ""))
        area = row.get("dien_tich_num", "")
        price = row.get("gia_ban_num", "")
        created_at = str(row.get("admin_created_at", ""))

        st.markdown(
            f"""
            <div class='admin-card {status_key}'>
                <div class='admin-title'>#{admin_post_id} • {title}</div>
                <div class='admin-meta'>📍 {quan} • 🏷 {loai_hinh} • 📐 {area} m² • 💰 {price} tỷ VNĐ</div>
                <div class='admin-meta'>🕒 Ngày đăng: {created_at}</div>
                <span class='admin-badge {status_key}'>{status_text_map.get(status_key, 'Chờ duyệt')}</span>
            </div>
            """,
            unsafe_allow_html=True,
        )

        c1, c2, c3 = st.columns([1, 1, 4])
        with c1:
            if st.button("✅ Duyệt bài", key=f"approve_{admin_post_id}"):
                update_admin_post_status(admin_post_id, "approved")
                st.rerun()
        with c2:
            if st.button("🗑️ Xóa bài", key=f"delete_{admin_post_id}"):
                update_admin_post_status(admin_post_id, "deleted")
                st.rerun()
        with c3:
            st.empty()

    csv = display_df.to_csv(index=False).encode("utf-8")
    st.download_button(
        "⬇️ Tải toàn bộ danh sách kiểm duyệt (CSV)",
        csv,
        file_name="admin_posts_status.csv",
        mime="text/csv",
    )


# =============================
# Main app
# =============================


def main() -> None:
    inject_global_css()

    st.title("🏙️ Hệ thống đăng tin & tìm kiếm nhà đất")

    if "page" not in st.session_state:
        st.session_state["page"] = "buyer"

    page = st.session_state["page"]

    menu = st.sidebar.selectbox("Chọn giao diện", ["Người mua", "Người bán", "Admin"])

    if "selected_house" not in st.session_state:
        st.session_state["selected_house"] = None
    if "selected_house_idx" not in st.session_state:
        st.session_state["selected_house_idx"] = None

    if page == "house_details":
        house_details_interface()
    elif menu == "Người bán":
        seller_interface()
    elif menu == "Admin":
        admin_interface()
    else:
        buyer_interface()

    render_app_footer()


if __name__ == "__main__":
    main()